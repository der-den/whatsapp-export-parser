#!/usr/bin/env python3

import os
import cv2
import subprocess
import re
import hashlib
import shutil
import json
import warnings
import torch
from pathlib import Path
from PIL import Image
from mutagen import File as MutagenFile
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from pptx import Presentation
from models import ContentType, ChatMessage
from utils import debug_print
from webp_handler import check_webp_animation, is_valid_sticker, extract_sticker_frames
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import whisper
import time

"""
Erstellt bzw Exrahiert für/von Attachments weitere Informationen und fügte diese als JSON Objekte der content variable hinzu. 
"""

class MetaParser:
    def __init__(self, zip_handler):
        self.zip_handler = zip_handler
        self.preview_success: Dict[ContentType, int] = {}
        self.attachment_counter = 0  # Initialize counter for attachments
        self.total_audio_files = 0  # Initialize total audio files counter
        self.current_audio_file = 0  # Initialize current audio file counter
        self.transcription_stats = {
            "transcoded": 0,
            "loaded_existing": 0,
            "errors": 0
        }
        
        # Load configuration
        self.config = self._load_config()
        
        # Initialize Whisper model if transcription is enabled
        self.model = None
        self.device = None
        if self.config["audio"]["transcription_enabled"]:
            self.device = "cuda" if torch.cuda.is_available() else "cpu"
            print(f"[Whisper] Using device: {self.device} for transcription")
            
            if torch.cuda.is_available():
                print(f"[Whisper] GPU device: {torch.cuda.get_device_name(0)}")
                print(f"[Whisper] CUDA version: {torch.version.cuda}")
                print(f"[Whisper] Available GPU memory: {torch.cuda.get_device_properties(0).total_memory / 1024**3:.2f} GB")
            
            # Get the model name from config
            model_name = self.config["audio"]["whisper_model"]
            
            # Load the model
            start_time = time.time()
            print(f"[Whisper] Loading model: {model_name}")
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=FutureWarning, 
                                     message='.*torch.load.*weights_only=False.*')
                self.model = whisper.load_model(model_name, device=self.device)
            
            load_time = time.time() - start_time
            model_size = sum(p.numel() for p in self.model.parameters()) / 1e6
            print(f"[Whisper] Model loaded in {load_time:.2f} seconds. Size: {model_size:.1f}M parameters")
        
    def _load_config(self) -> dict:
        """Load configuration from config.json"""
        config_path = os.path.join(os.path.dirname(__file__), 'config.json')
        try:
            with open(config_path, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            # Return default configuration if file not found
            return {
                "audio": {
                    "transcription_enabled": True,
                    "whisper_model": "large"
                },
                "output": {
                    "include_attachments": True,
                    "max_image_width": 800,
                    "max_image_height": 600
                }
            }

    def _calculate_md5(self, file_path: str) -> str:
        """
        Berechnet den MD5-Hash einer Datei.
        
        Args:
            file_path: Pfad zur Datei
            
        Returns:
            str: MD5-Hash der Datei
        """
        try:
            with open(file_path, 'rb') as f:
                file_hash = hashlib.md5()
                while chunk := f.read(8192):
                    file_hash.update(chunk)
            return file_hash.hexdigest()
        except Exception as e:
            debug_print(f"Error calculating MD5 hash: {e}", component="meta")
            return ""

    def _get_meta_directory(self) -> str:
        """
        Erstellt und gibt das Meta-Verzeichnis zurück, parallel zum Extraktionsverzeichnis.
        """
        # Extrahiere den Hash-Namen aus dem Extraktionspfad
        extract_dir_name = os.path.basename(self.zip_handler.extract_path)
        meta_dir = os.path.join(os.path.dirname(self.zip_handler.extract_path), f"{extract_dir_name}_meta")
        transcribe_dir = os.path.join(meta_dir, "transcribe")
        stickerframes_dir = os.path.join(meta_dir, "stickerframes")
        
        # Create directories if they don't exist
        os.makedirs(meta_dir, exist_ok=True)
        os.makedirs(transcribe_dir, exist_ok=True)
        os.makedirs(stickerframes_dir, exist_ok=True)
        
        debug_print(f"Creating meta directory: {meta_dir}", component="meta")

        return meta_dir

    def _transcribe_audio(self, file_path: str, audio_file: str) -> Dict:
        """
        Transcribes an audio file using Whisper and returns the transcription metadata.
        
        Args:
            file_path: Path to the audio file
            audio_file: Name of the audio file
            
        Returns:
            dict: Transcription metadata including the text, model info, and any error information
        """
        result = {
            "success": False,
            "error": None,
            "error_type": None
        }
        
        # Check if transcription is enabled in config
        if not self.config["audio"]["transcription_enabled"]:
            result["error"] = "Audio transcription is disabled in configuration"
            result["error_type"] = "transcription_disabled"
            return result

        # Check for existing transcription
        meta_dir = self._get_meta_directory()
        model_name = self.config["audio"]["whisper_model"]
        
        # Create a more detailed filename with model info and attachment ID
        transcribe_dir = os.path.join(meta_dir, "transcribe")
        os.makedirs(transcribe_dir, exist_ok=True)
        
        # Extract model info
        model_info = f"whisper-{model_name}"
        
        # Create filename with attachment number
        base_name = os.path.splitext(audio_file)[0]
        json_filename = f"{base_name}.att{self.attachment_counter}.{model_info}.json"
        json_output = os.path.join(transcribe_dir, json_filename)
        
        # Check if transcription file already exists
        if os.path.exists(json_output):
            print(f"#{self.attachment_counter} [Whisper] Found existing transcription for {audio_file}", end="\r")      
            debug_print(f"Loading existing transcription for #{self.attachment_counter} - {audio_file}", component="meta")
            with open(json_output, 'r', encoding='utf-8') as f:
                existing_result = json.load(f)
                if "transcription" in existing_result:
                    text = existing_result["transcription"].get("text", "")
                    model = existing_result["transcription"].get("model", f"whisper-{model_name}")
                    language = existing_result["transcription"].get("language", "unknown")
                    result = {
                        "success": True,
                        "transcription": {
                            "text": text,
                            "model": model,
                            "language": language,
                            "transcribed_at": existing_result["transcription"].get("transcribed_at", datetime.now().isoformat())
                        }
                    }
                    self.transcription_stats["loaded_existing"] += 1
                    return result
        
        self.current_audio_file += 1
        debug_print(f"Transcribing audio: {audio_file} ({self.current_audio_file}/{self.total_audio_files})", component="meta")

        try:
            # Transcribe audio with GPU acceleration if available
            start_time = time.time()
            transcribe_result = self.model.transcribe(
                file_path,
                fp16=torch.cuda.is_available()  # Enable FP16 if CUDA is available
            )
            transcribe_time = time.time() - start_time
            # print transcription information, but reset line feed to show next print on same line
            print(f"#{self.attachment_counter} [Whisper] Transcription completed in {transcribe_time:.2f} seconds for {audio_file}", end="\r")

            if not transcribe_result or "text" not in transcribe_result:
                debug_print(f"Warning: No transcription result for {audio_file}", component="meta")
                result["error"] = "No transcription result"
                result["error_type"] = "empty_result"
                self.transcription_stats["errors"] += 1
                return result
            
            # Prepare metadata
            transcription_meta = {
                "text": transcribe_result["text"],
                "model": f"whisper-{model_name}",
                "language": transcribe_result.get("language", "unknown"),
                "segments": transcribe_result.get("segments", []),
                "transcribed_at": datetime.now().isoformat(),
                "device_used": self.device
            }
            
            result["transcription"] = transcription_meta
            result["success"] = True
            self.transcription_stats["transcoded"] += 1
            
            # Save transcription to file
            with open(json_output, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            
            return result
            
        except RuntimeError as e:
            error_msg = f"CUDA/GPU error while transcribing {audio_file}: {str(e)}"
            debug_print(error_msg, component="meta")
            result["error"] = error_msg
            result["error_type"] = "cuda_error"
            self.transcription_stats["errors"] += 1
            return result
        except Exception as e:
            error_msg = f"Error transcribing {audio_file}: {str(e)}"
            debug_print(error_msg, component="meta")
            result["error"] = error_msg
            result["error_type"] = "general_error"
            self.transcription_stats["errors"] += 1
            return result

    def _get_audio_metadata(self, audio_file: str) -> dict:
        """
        Extrahiert Metadaten von einer Audiodatei.
        
        Args:
            audio_file: Pfad zur Audiodatei
            
        Returns:
            dict: Metadaten der Audiodatei
        """
        try:
            file_path = os.path.join(self.zip_handler.extract_path, audio_file)
            audio = MutagenFile(file_path)
            
            self.attachment_counter += 1  # Increment counter
            self.total_audio_files += 1  # Increment total audio files counter
            
            metadata = {
                "type": "audio",
                "filename": audio_file,
                "attachment_number": self.attachment_counter,
                "format": audio.mime[0].split('/')[-1] if audio.mime else None,
                "duration_seconds": audio.info.length if hasattr(audio.info, 'length') else None,
                "channels": audio.info.channels if hasattr(audio.info, 'channels') else None,
                "size_bytes": os.path.getsize(file_path),
                "md5_hash": self._calculate_md5(file_path)
            }
            
            # Try to add transcription if available, but don't fail if it's not possible
            try:
                transcribe_result = self._transcribe_audio(file_path, audio_file)
                if transcribe_result["success"] and transcribe_result["transcription"]:
                    metadata["transcription"] = transcribe_result["transcription"]
                elif transcribe_result["error"]:
                    metadata["transcription_error"] = {
                        "error": transcribe_result["error"],
                        "error_type": transcribe_result["error_type"]
                    }
            except Exception as e:
                debug_print(f"Error during transcription: {e}", component="meta")
            
            # Remove None values
            metadata = {k: v for k, v in metadata.items() if v is not None}
            
            return metadata
        except Exception as e:
            debug_print(f"Error extracting audio metadata: {e}", component="meta")
            return {"error": str(e)}

    def _get_image_metadata(self, image_file: str) -> dict:
        """
        Extrahiert Metadaten von einem Bild.
        
        Args:
            image_file: Name der Bilddatei
            
        Returns:
            dict: Metadaten des Bildes
        """
        try:
            file_path = os.path.join(self.zip_handler.extract_path, image_file)
            with Image.open(file_path) as img:
                self.attachment_counter += 1  # Increment counter
                metadata = {
                    "type": "image",
                    "filename": image_file,
                    "attachment_number": self.attachment_counter,
                    "width": img.width,
                    "height": img.height,
                    "format": img.format,
                    "mode": img.mode,
                    "size_bytes": os.path.getsize(file_path),
                    "md5_hash": self._calculate_md5(file_path)
                }
                return metadata
        except Exception as e:
            debug_print(f"Error extracting image metadata: {e}", component="meta")
            return {"error": str(e)}

    def _get_video_metadata(self, video_file: str, preview_paths: Optional[Tuple[str, str]] = None) -> dict:
        """
        Extrahiert Metadaten von einem Video.
        
        Args:
            video_file: Pfad zum Video
            preview_paths: Optional tuple mit meta_path und report_path für Preview
            
        Returns:
            dict: Metadaten des Videos
        """
        try:
            file_path = os.path.join(self.zip_handler.extract_path, video_file)
            cap = cv2.VideoCapture(file_path)
            self.attachment_counter += 1  # Increment counter
            metadata = {
                "type": "video",
                "filename": video_file,
                "attachment_number": self.attachment_counter,
                "width": int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)),
                "height": int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)),
                "frame_count": int(cap.get(cv2.CAP_PROP_FRAME_COUNT)),
                "fps": float(cap.get(cv2.CAP_PROP_FPS)),
                "duration_seconds": float(cap.get(cv2.CAP_PROP_FRAME_COUNT) / cap.get(cv2.CAP_PROP_FPS)),
                "size_bytes": os.path.getsize(file_path),
                "md5_hash": self._calculate_md5(file_path)
            }
            
            if preview_paths:
                meta_path, report_path = preview_paths
                metadata["preview"] = {
                    "meta_path": meta_path,
                    "report_path": report_path
                }
                
            cap.release()
            return metadata
        except Exception as e:
            debug_print(f"Error extracting video metadata: {e}", component="meta")
            return {"error": str(e)}

    def _get_document_metadata(self, doc_file: str) -> dict:
        """
        Extrahiert Metadaten von einem Dokument (PDF, DOCX, XLSX, PPTX).
        
        Args:
            doc_file: Pfad zum Dokument
            
        Returns:
            dict: Metadaten des Dokuments
        """
        try:
            file_path = os.path.join(self.zip_handler.extract_path, doc_file)
            file_ext = os.path.splitext(doc_file)[1].lower()
            
            self.attachment_counter += 1  # Increment counter
            metadata = {
                "type": "document",
                "filename": doc_file,
                "attachment_number": self.attachment_counter,
                "format": file_ext[1:],  # Remove the dot
                "size_bytes": os.path.getsize(file_path),
                "md5_hash": self._calculate_md5(file_path)
            }
            
            # PDF-spezifische Metadaten
            if file_ext == '.pdf':
                with open(file_path, 'rb') as file:
                    pdf = PdfReader(file)
                    info = pdf.metadata
                    if info:
                        metadata.update({
                            "title": info.get('/Title', None),
                            "author": info.get('/Author', None),
                            "page_count": len(pdf.pages)
                        })
            
            # DOCX-spezifische Metadaten
            elif file_ext == '.docx':
                doc = Document(file_path)
                core_properties = doc.core_properties
                metadata.update({
                    "title": core_properties.title,
                    "author": core_properties.author,
                    "page_count": len(doc.sections)
                })
            
            # XLSX-spezifische Metadaten
            elif file_ext == '.xlsx':
                wb = load_workbook(file_path, read_only=True)
                metadata.update({
                    "title": wb.properties.title,
                    "author": wb.properties.creator,
                    "sheet_count": len(wb.sheetnames)
                })
                wb.close()
            
            # PPTX-spezifische Metadaten
            elif file_ext == '.pptx':
                prs = Presentation(file_path)
                core_properties = prs.core_properties
                metadata.update({
                    "title": core_properties.title,
                    "author": core_properties.author,
                    "slide_count": len(prs.slides)
                })
            
            # Remove None values
            metadata = {k: v for k, v in metadata.items() if v is not None}
            
            return metadata
        except Exception as e:
            debug_print(f"Error extracting document metadata: {e}", component="meta")
            return {"error": str(e)}

    def _get_sticker_metadata(self, sticker_file: str) -> dict:
        """
        Extracts metadata from a sticker file (WebP).
        
        Args:
            sticker_file: Path to the sticker file
            
        Returns:
            dict: Metadata of the sticker file
        """
        try:
            file_path = os.path.join(self.zip_handler.extract_path, sticker_file)
            self.attachment_counter += 1  # Increment counter
            metadata = {
                "type": "sticker",
                "filename": sticker_file,
                "attachment_number": self.attachment_counter,
                "format": "webp",
                "size_bytes": os.path.getsize(file_path),
                "md5_hash": self._calculate_md5(file_path)
            }
            return metadata
        except Exception as e:
            debug_print(f"Error extracting sticker metadata: {e}", component="meta")
            return {"error": str(e)}

    def _take_video_frames(self, video_file: str) -> Optional[Tuple[str, str]]:
        """
        Extrahiert 4 Frames aus einem Video (bei 20%, 40%, 60% und 80% der Gesamtlänge)
        und fügt sie zu einem Vorschaubild zusammen.
        
        Args:
            video_file: Der Dateiname des Videos
            
        Returns:
            Tuple[str, str]: (Relativer Pfad zum Bild im Meta-Dir, Relativer Pfad im Report) oder None bei Fehler
        """
        try:
            # Erstelle Meta-Verzeichnis und Unterverzeichnisse
            meta_dir = self._get_meta_directory()
            frames_dir = os.path.join(meta_dir, 'videoframes')
            os.makedirs(frames_dir, exist_ok=True)
            
            # Erstelle MD5-Hash des Videonamens für den Dateinamen
            video_hash = hashlib.md5(video_file.encode()).hexdigest()
            frame_file = f"{video_hash}.png"
            frame_path = os.path.join(frames_dir, frame_file)
            
            # Pfade für den Report
            report_images_dir = os.path.join(self.zip_handler.extract_path, 'html', 'images', 'videoframes')
            os.makedirs(report_images_dir, exist_ok=True)
            report_frame_path = os.path.join(report_images_dir, frame_file)
            
            # Wenn Frame bereits existiert
            if os.path.exists(frame_path):
                # Kopiere zum Report falls noch nicht vorhanden
                if not os.path.exists(report_frame_path):
                    shutil.copy2(frame_path, report_frame_path)
                return os.path.join('videoframes', frame_file), os.path.join('images', 'videoframes', frame_file)
            
            # Öffne das Video
            video_path = os.path.join(self.zip_handler.extract_path, video_file)
            cap = cv2.VideoCapture(video_path)
            
            if not cap.isOpened():
                debug_print(f"Error: Could not open video {video_file}", component="meta")
                return None
            
            # Hole Video-Informationen
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            if total_frames == 0:
                debug_print(f"Error: Video {video_file} has no frames", component="meta")
                return None
            
            # Berechne Frame-Positionen (20%, 40%, 60%, 80%)
            frame_positions = [int(total_frames * pos) for pos in [0.2, 0.4, 0.6, 0.8]]
            frames = []
            
            # Extrahiere die Frames
            for pos in frame_positions:
                cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
                ret, frame = cap.read()
                if ret:
                    # Konvertiere BGR zu RGB
                    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    frames.append(frame_rgb)
                else:
                    debug_print(f"Error taking video frame: {str(e)}", component="meta")
            
            cap.release()
            
            if len(frames) != 4:
                debug_print(f"Error taking video frames: {str(e)}", component="meta")
                return None
            
            # Skaliere die Frames auf einheitliche Größe
            target_size = (320, 180)  # 16:9 Format
            scaled_frames = []
            for frame in frames:
                pil_img = Image.fromarray(frame)
                pil_img.thumbnail(target_size, Image.Resampling.LANCZOS)
                # Erstelle neues Bild mit weißem Hintergrund
                new_img = Image.new('RGB', target_size, (255, 255, 255))
                # Zentriere das Bild
                x = (target_size[0] - pil_img.size[0]) // 2
                y = (target_size[1] - pil_img.size[1]) // 2
                new_img.paste(pil_img, (x, y))
                scaled_frames.append(new_img)
            
            # Erstelle 2x2 Grid
            grid_size = (target_size[0] * 2, target_size[1] * 2)
            grid_image = Image.new('RGB', grid_size, (255, 255, 255))
            
            # Füge Frames zum Grid hinzu
            for i, frame in enumerate(scaled_frames):
                x = (i % 2) * target_size[0]
                y = (i // 2) * target_size[1]
                grid_image.paste(frame, (x, y))
            
            # Speichere das finale Bild
            grid_image.save(frame_path, 'PNG')
            
            # Kopiere zum Report
            shutil.copy2(frame_path, report_frame_path)
            
            # Logge Video und Frame-Name
            log_file = os.path.join(meta_dir, 'videoframes.log')
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(f"{video_file}\t{frame_file}\n")
            
            return os.path.join('videoframes', frame_file), os.path.join('images', 'videoframes', frame_file)
            
        except Exception as e:
            debug_print(f"Error taking video frames: {str(e)}", component="meta")
            return None

    def _take_webpage_screenshot(self, url: str) -> Optional[Tuple[str, str]]:
        """
        Macht einen Screenshot einer Webseite und speichert ihn als PNG.
        
        Args:
            url: Die URL der Webseite
            
        Returns:
            Tuple[str, str]: (Relativer Pfad zum Screenshot im Meta-Dir, Relativer Pfad im Report) oder None bei Fehler
        """
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            
            # Erstelle Meta-Verzeichnis und Unterverzeichnisse
            meta_dir = self._get_meta_directory()
            screenshots_dir = os.path.join(meta_dir, 'linkshots')
            os.makedirs(screenshots_dir, exist_ok=True)
            
            # Erstelle MD5-Hash der URL für den Dateinamen
            url_hash = hashlib.md5(url.encode()).hexdigest()
            screenshot_file = f"{url_hash}.png"
            screenshot_path = os.path.join(screenshots_dir, screenshot_file)
            
            # Pfade für den Report
            report_images_dir = os.path.join(self.zip_handler.extract_path, 'html', 'images', 'screenshots')
            os.makedirs(report_images_dir, exist_ok=True)
            report_screenshot_path = os.path.join(report_images_dir, screenshot_file)
            
            # Wenn Screenshot bereits existiert
            if os.path.exists(screenshot_path):
                # Kopiere zum Report falls noch nicht vorhanden
                if not os.path.exists(report_screenshot_path):
                    shutil.copy2(screenshot_path, report_screenshot_path)
                return os.path.join('linkshots', screenshot_file), os.path.join('images', 'screenshots', screenshot_file)
            
            # Chrome Optionen setzen
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--window-size=1920,1080')
            
            # Starte Chrome
            driver = webdriver.Chrome(options=chrome_options)
            
            try:
                # Lade die Seite
                driver.get(url)
                driver.implicitly_wait(5)
                
                # Mache den Screenshot
                driver.save_screenshot(screenshot_path)
                debug_print(f"Taking screenshot of: {url}", component="meta")
                
                # Kopiere Screenshot in den Report
                shutil.copy2(screenshot_path, report_screenshot_path)
                
                # Logge URL und Screenshot-Name
                log_file = os.path.join(meta_dir, 'linkshots.log')
                with open(log_file, 'a', encoding='utf-8') as f:
                    f.write(f"{url}\t{screenshot_file}\n")
                
                return os.path.join('linkshots', screenshot_file), os.path.join('images', 'screenshots', screenshot_file)
                
            finally:
                driver.quit()
                
        except Exception as e:
            debug_print(f"Error taking screenshot: {str(e)}", component="meta")
            return None

    def process_messages(self, messages: List[ChatMessage], url_pattern: str) -> Dict[ContentType, int]:
        """
        Verarbeitet eine Liste von Nachrichten, extrahiert Meta-Informationen und speichert sie als JSON im Content.
        
        Args:
            messages: Liste der Chat-Nachrichten
            url_pattern: Regex-Pattern zum Finden von URLs
            
        Returns:
            Dict[ContentType, int]: Anzahl der erfolgreichen Previews pro ContentType
        """
        total_messages = len(messages)

        for i, message in enumerate(messages):
            
            # show progress
            if i % 100 == 0 or i == total_messages - 1:
                progress = (i / total_messages) * 100
                debug_print(f"Progress: {progress:.1f}% ({i+1}/{total_messages} messages)", component="meta")
            
            if message.is_attachment:
                debug_print(f"Current Attachment Number: {self.attachment_counter}", component="meta")

            # Process Images
            if message.content_type.is_image and message.attachment_file:
                debug_print(f"Processing image: {message.attachment_file}", component="meta")
                try:
                    metadata = self._get_image_metadata(message.attachment_file)
                    if metadata and "error" not in metadata:
                        message.content = json.dumps(metadata, ensure_ascii=False)
                        debug_print(f"Added metadata for image: {message.attachment_file}", component="meta")
                except Exception as e:
                    debug_print(f"Error processing image: {e}", component="meta")
                self.preview_success[ContentType.IMAGE] = self.preview_success.get(ContentType.IMAGE, 0) + 1
        
            # Process Videos
            if message.content_type.is_video and message.attachment_file:
                debug_print(f"Taking video frames for: {message.attachment_file}", component="meta")
                try:
                    frame_paths = self._take_video_frames(message.attachment_file)
                    metadata = self._get_video_metadata(message.attachment_file, frame_paths)
                    if metadata and "error" not in metadata:
                        message.content = json.dumps(metadata, ensure_ascii=False)
                        debug_print(f"Added metadata and preview for video: {message.attachment_file}", component="meta")
                except Exception as e:
                    debug_print(f"Error taking video frames: {str(e)}", component="meta")      
                self.preview_success[ContentType.VIDEO] = self.preview_success.get(ContentType.VIDEO, 0) + 1
        
            # Process Audio
            if message.content_type.is_audio and message.attachment_file:
                debug_print(f"Processing audio: {message.attachment_file}", component="meta")
                try:
                    metadata = self._get_audio_metadata(message.attachment_file)
                    if metadata and "error" not in metadata:
                        message.content = json.dumps(metadata, ensure_ascii=False)
                        debug_print(f"Added metadata for audio: {message.attachment_file}", component="meta")
                except Exception as e:
                    debug_print(f"Error processing audio: {e}", component="meta")
                self.preview_success[ContentType.AUDIO] = self.preview_success.get(ContentType.AUDIO, 0) + 1
        
            # Process Documents
            if message.content_type.is_document and message.attachment_file:
                debug_print(f"Processing document: {message.attachment_file}", component="meta")
                try:
                    metadata = self._get_document_metadata(message.attachment_file)
                    if metadata and "error" not in metadata:
                        message.content = json.dumps(metadata, ensure_ascii=False)
                        debug_print(f"Added metadata for document: {message.attachment_file}", component="meta")
                except Exception as e:
                    debug_print(f"Error processing document: {e}", component="meta")
                self.preview_success[ContentType.DOCUMENT] = self.preview_success.get(ContentType.DOCUMENT, 0) + 1

            # Process Stickers/WebP
            if message.content_type == ContentType.STICKER and message.attachment_file:
                debug_print(f"Processing sticker: {message.attachment_file}", component="meta")
                try:
                    metadata = self._get_sticker_metadata(message.attachment_file)
                    if metadata and "error" not in metadata:
                        # If sticker is multiframe, extract frames
                        if message.is_multiframe:
                            sticker_path = os.path.join(self.zip_handler.extract_path, message.attachment_file)
                            frames_dir = os.path.join(self._get_meta_directory(), "stickerframes", f"sticker_{metadata['attachment_number']}")
                            os.makedirs(frames_dir, exist_ok=True)
                            frame_paths = extract_sticker_frames(sticker_path, frames_dir)
                            
                            # Add frame info to metadata
                            metadata["frames"] = {
                                "count": len(frame_paths),
                                "paths": [os.path.relpath(p, self._get_meta_directory()) for p in frame_paths]
                            }
                         # if frame count = 1, set is_multiframe to False
                        if metadata["frames"]["count"] == 1:
                            metadata["is_multiframe"] = False

                        message.content = json.dumps(metadata, ensure_ascii=False)
                        debug_print(f"Added metadata for sticker: {message.attachment_file}", component="meta")
                except Exception as e:
                    debug_print(f"Error processing sticker: {e}", component="meta")
                self.preview_success[ContentType.STICKER] = self.preview_success.get(ContentType.STICKER, 0) + 1

            # Process Links
            if message.content_type == ContentType.LINK:
                try:
                    urls = re.findall(url_pattern, message.content)
                    if urls:
                        url = urls[0]
                        debug_print(f"Found URL: {url}", component="meta")
                        # Screenshot deaktiviert - Code bleibt für spätere Verwendung
                        '''
                        screenshot_paths = self._take_webpage_screenshot(url)
                        if screenshot_paths:
                            meta_path, report_path = screenshot_paths
                            message.content = f"{message.content}\n[Screenshot: {meta_path}]"
                            message.content += f'\n<div class="link-preview"><img src="{report_path}" alt="Screenshot of {url}" class="link-screenshot"/></div>'
                            debug_print(f"Created screenshot for {url}", component="meta")
                            self.preview_success[ContentType.LINK] = self.preview_success.get(ContentType.LINK, 0) + 1
                        else:
                            debug_print(f"Failed to create screenshot for {url}", component="meta")
                        '''
                except Exception as e:
                    debug_print(f"Error processing URL: {e}", component="meta")
                self.preview_success[ContentType.LINK] = self.preview_success.get(ContentType.LINK, 0) + 1
            

        debug_print("\n\nProcessing Summary:", component="meta")
        for content_type, count in self.preview_success.items():
            debug_print(f"  {content_type.name}: {count}", component="meta")
        
        if self.transcription_stats["transcoded"] > 0 or self.transcription_stats["errors"] > 0:
            debug_print(f"\nAudio Processing Summary:", component="meta")
            debug_print(f"Total audio files processed: {self.preview_success[ContentType.AUDIO]}", component="meta")
            debug_print(f"- Successfully transcribed: {self.transcription_stats['transcoded']}", component="meta")
            debug_print(f"- Loaded from existing: {self.transcription_stats['loaded_existing']}", component="meta")
            debug_print(f"- Failed transcriptions: {self.transcription_stats['errors']}", component="meta")

        return self.preview_success
